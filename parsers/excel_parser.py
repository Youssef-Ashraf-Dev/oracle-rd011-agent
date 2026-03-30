"""
RD.011 Agent — Excel (.xlsx) parser.

Extracts structured text from Oracle Finance questionnaire workbooks.
Handles merged cells, multi-sheet workbooks, and blank cells gracefully.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

logger = logging.getLogger(__name__)


def _resolve_merged_value(ws, row: int, col: int):
    """
    If (row, col) falls inside a merged range, return the value of the
    top-left cell of that range.  Otherwise return the cell's own value.
    """
    for rng in ws.merged_cells.ranges:
        if (row, col) in [(r, c) for r in range(rng.min_row, rng.max_row + 1)
                          for c in range(rng.min_col, rng.max_col + 1)]:
            return ws.cell(row=rng.min_row, column=rng.min_col).value
    return ws.cell(row=row, column=col).value


def _sheet_to_text(ws) -> str:
    """Convert a single worksheet to structured text."""
    if ws.max_row is None or ws.max_row == 0:
        return ""

    max_row = ws.max_row
    max_col = ws.max_column or 1

    # Gather all rows, resolving merged cells
    rows_raw: list[list[str]] = []
    for row_idx in range(1, max_row + 1):
        row_values: list[str] = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                val = _resolve_merged_value(ws, row_idx, col_idx)
            else:
                val = cell.value
            row_values.append(str(val).strip() if val is not None else "")
        rows_raw.append(row_values)

    # Drop completely empty rows
    rows_data = [row for row in rows_raw if any(row)]
    if not rows_data:
        return ""

    # Drop completely empty columns
    num_cols = max(len(row) for row in rows_data)
    keep_cols = [
        col_idx for col_idx in range(num_cols)
        if any(row[col_idx] for row in rows_data if col_idx < len(row))
    ]
    if not keep_cols:
        return ""

    filtered = [
        [row[i] for i in keep_cols if i < len(row)]
        for row in rows_data
    ]

    # Format as markdown table; first non-empty row is the header
    header = filtered[0]
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in filtered[1:]:
        padded = row[:len(header)]
        while len(padded) < len(header):
            padded.append("")
        lines.append("| " + " | ".join(padded) + " |")

    return "\n".join(lines)


def parse_excel(file_path: str) -> str:
    """
    Parse an ``.xlsx`` file into a structured text string.

    Uses ``openpyxl``.  For each sheet:

    - Sheet name as heading
    - Reads all non-empty cells including merged cell values
    - Formats as markdown table
    - Handles blank cells gracefully (writes empty string)

    Parameters
    ----------
    file_path
        Path to the ``.xlsx`` file.

    Returns
    -------
    str
        All sheets as a combined string.
    """
    path = Path(file_path)
    if not path.suffix.lower() == ".xlsx":
        raise ValueError(f"Expected .xlsx file, got: {path.suffix}")
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    logger.info("Parsing Excel: %s", file_path)
    wb = load_workbook(str(path), read_only=False, data_only=True)

    output_parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_text = _sheet_to_text(ws)
        if sheet_text:
            output_parts.append(f"## {sheet_name}\n\n{sheet_text}")

    wb.close()

    result = "\n\n".join(output_parts)
    logger.info("Parsed %d characters from %s", len(result), path.name)
    return result
